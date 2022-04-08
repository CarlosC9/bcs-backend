import json
import logging
import os
import subprocess
import traceback

import sys
from enum import Enum
from typing import Dict, List, Callable, Union, Optional
from urllib.parse import unquote

import redis
from alchemyjsonschema import SchemaFactory, StructuralWalker
from attr import attrs, attrib
from flask import Response, Blueprint, g, request
from flask.views import MethodView
import sqlalchemy
from sqlalchemy import orm, and_, or_
from sqlalchemy.orm import Query
from sqlalchemy.pool import StaticPool
from bioblend import galaxy

from .. import config_file_var, app_acronym
from ..authentication import n_session
from ..common import generate_json, ROOT
from ..common.pg_helpers import create_pg_database_engine, load_table, load_many_to_many_table, \
    load_computing_resources, load_processes_in_computing_resources, load_process_input_schema, load_table_extended
from ..db_models import DBSession, DBSessionChado, ORMBaseChado, DBSessionGeo
from ..db_models.bioinformatics import *
from ..db_models.core import update_functional_object_tsvector
from ..db_models.geographics import *
from ..db_models.hierarchies import HierarchyType, Hierarchy, HierarchyNode
from ..db_models.jobs import *
from ..db_models.sa_annotations import *
from ..db_models.sysadmin import *
from ..rest.socket_service import SocketService
import biobarcoding as base_app_pkg

app_api_base = "/api"  # Base for all RESTful calls
app_gui_base = "/gui"  # Base for the Angular2 GUI
app_external_gui_base = "/gui_external"  # Base for the Angular2 GUI when loaded from URL
app_proxy_base = "/pxy"  # Base for the Backend Proxy

log_level = logging.DEBUG
logger = logging.getLogger('app_logger')

logger.setLevel(log_level)  # set logger level
logFormatter = logging.Formatter("%(name)-12s %(asctime)s %(levelname)-8s %(filename)s:%(funcName)s %(message)s")
consoleHandler = logging.StreamHandler(sys.stdout)  # set streamhandler to stdout
consoleHandler.setFormatter(logFormatter)
logger.addHandler(consoleHandler)


def build_json_response(obj, status=200):
    return Response(generate_json(obj),
                    mimetype="text/json",
                    status=status)


class IType(Enum):
    INFO = 1
    WARNING = 2
    ERROR = 3


@attrs
class IssueLocation:
    # TODO IssueLocation
    def __str__(self):
        return f"Put your Ad here :)"


@attrs
class Issue:
    itype = attrib()
    message = attrib()
    location = attrib(default=None)

    def as_dict(self):
        return {"type": self.itype.name, "message": self.message, "location": self.location}

    def __str__(self):
        return f'(type="{self.itype.name}", message="{self.message}", location="{self.location}")'


class ResponseObject:
    def __init__(self, content=None, count=0, issues=None, content_type="text/json", status=200):
        self.content = content
        self.count = count
        self.issues = [] if issues is None else issues
        self.content_type = content_type
        self.status = status

    def get_response(self) -> Response:
        """
        status == 200. ctype not JSON -> content, ctype, 200
                       ctype     JSON -> dict(issues=issues, content=content), ctype, 200
        status != 200. ctype not JSON -> dict(issues=issues), "text/json", status
                       ctype     JSON -> dict(issues=issues)
        :return:
        """
        if self.status == 200:
            if self.content_type in ("text/json", "application/json"):
                obj = generate_json(dict(issues=[s.as_dict() for s in self.issues],
                                         content=self.content, count=self.count))
            else:
                obj = self.content
        else:
            self.content_type = "application/json"
            obj = generate_json(dict(issues=[s.as_dict() for s in self.issues]))

        return Response(obj, mimetype=self.content_type, status=self.status)


def get_default_configuration_dict():
    from appdirs import AppDirs
    app_dirs = AppDirs(f"{app_acronym}-backend")
    # Default directories, multi-platform
    data_path = app_dirs.user_data_dir
    cache_path = app_dirs.user_cache_dir

    REDIS_HOST = "redis"
    REDIS_PORT = 6379
    BROKER_URL = f"redis://{REDIS_HOST}:{REDIS_PORT}/0"
    BACKEND_URL = BROKER_URL

    return dict(
        # SYSTEM DB
        DB_CONNECTION_STRING="postgresql://postgres:postgres@localhost:5432/",
        # CHADO (MOLECULAR DATA DB)
        CHADO_DATABASE="postgres",
        CHADO_HOST="localhost",
        CHADO_PORT="5432",
        CHADO_USER="postgres",
        CHADO_PASSWORD="postgres",
        CHADO_SCHEMA="public",
        # CELERY (TASK EXECUTION)
        CELERY_BROKER_URL=f"{BROKER_URL}",
        CELERY_BACKEND_URL=f"{BACKEND_URL}",
        REDIS_HOST="filesystem:local_session",
        REDIS_HOST_FILESYSTEM_DIR=f"{cache_path}/sessions",
        # GEO (GEOSPATIAL DATA)
        GEOSERVER_USER="admin",
        GEOSERVER_PASSWORD=f"{app_acronym}_ad37",
        GEOSERVER_HOST="localhost",
        GEOSERVER_PORT="9180",
        # GEOSERVER address from App-Backend, used by the GeoserverProxy
        GEOSERVER_URL="localhost:9180",
        # PostGIS address from App-Backend
        POSTGIS_CONNECTION_STRING="postgresql://postgres:postgres@localhost:5435/",
        # PostGIS address from Geoserver ("host", "port" could differ from those in POSTGIS_CONNECTION_STRING)
        POSTGIS_USER="postgres",
        POSTGIS_PASSWORD="postgres",
        POSTGIS_PORT="5432",
        POSTGIS_HOST="localhost",
        POSTGIS_DB=f"{app_acronym}_geoserver",
        # COMPUTE RESOURCES
        RESOURCES_CONFIG_FILE_PATH=f"{data_path}/compute_resources.json",
        JOBS_LOCAL_WORKSPACE=os.path.expanduser(f'~/{app_acronym}_jobs'),
        SSH_JOBS_DEFAULT_REMOTE_WORKSPACE="/tmp",
        GALAXY_API_KEY="fakekey",
        GALAXY_LOCATION="http://localhost:8080",
        # MISC
        COLOR_PALETTES=f"{data_path}/color_palettes.yaml",
        GOOGLE_APPLICATION_CREDENTIALS=f"{data_path}/firebase-key.json",  # Firebase
        ENDPOINT_URL="http://localhost:5000",  # Self "app-backend" address (so Celery can update things)
        COOKIES_FILE_PATH=f"/tmp/{app_acronym}-cookies.txt",  # Where cookies are stored by Celery
        CACHE_FILE_LOCATION=f"{cache_path}/cache",  # Cached things
        SAMESITE_NONE="True",
        TESTING="True",
        SELF_SCHEMA="",
        ACL_ENABLED="False",
    )


def prepare_default_configuration(create_directories):
    default_cfg = get_default_configuration_dict()

    from appdirs import AppDirs
    app_dirs = AppDirs(f"{app_acronym}-backend")

    # Default directories, multi-platform
    data_path = app_dirs.user_data_dir
    cache_path = app_dirs.user_cache_dir
    if create_directories:
        os.makedirs(data_path, exist_ok=True)
        os.makedirs(cache_path, exist_ok=True)
        directories = [v for k, v in default_cfg.items() if k in ["CACHE_FILE_LOCATION", "REDIS_HOST_FILESYSTEM_DIR"]]
        # Obtain and create directories
        for v in directories:
            os.makedirs(v, exist_ok=True)

    # Default configuration
    return f"""{os.linesep.join([f'{k}="{v}"' for k, v in default_cfg.items()])}""", \
           data_path + os.sep + f"{app_acronym}_local.conf"


def complete_configuration_file(file_name):
    default_cfg = get_default_configuration_dict()
    default_cfg_keys = list(default_cfg.keys())
    with open(file_name, 'r+') as file:
        constants = file.read().splitlines()
        for const in constants:
            if const and not const.strip().startswith('#'):
                key, _ = const.split("=")
                if key.strip() in default_cfg_keys:
                    default_cfg_keys.remove(key.strip())
                else:
                    print(f"WARNING: {key.strip()} is in the configuration file but not in the default config")
        for key in default_cfg_keys:
            print(f"{key} inserted to {file_name} with value {default_cfg[key]}")
            file.write(f'\n{key}="{default_cfg[key]}"')


def load_configuration_file(flask_app):
    # Initialize configuration
    try:
        _, file_name = prepare_default_configuration(False)
        if not os.environ.get(config_file_var):
            logger.debug(f"Trying {file_name}")
            if os.path.isfile(file_name):
                print(f"Assuming {file_name} as configuration file")
                logger.debug(f"Assuming {file_name} as configuration file")
                found = True
                complete_configuration_file(file_name)
                os.environ[config_file_var] = file_name
            else:
                found = False
                logger.debug(f"Creating {file_name} as configuration file")
            if not found:
                cfg, file_name = prepare_default_configuration(True)
                print(f"Generating {file_name} as configuration file:\n{cfg}")
                with open(file_name, "wt") as f:
                    f.write(cfg)
                os.environ[config_file_var] = file_name
        else:
            complete_configuration_file(os.environ.get(config_file_var))
        print("-----------------------------------------------")
        print(f'Configuration file at: {os.environ[config_file_var]}')
        print("-----------------------------------------------")
        flask_app.config.from_envvar(config_file_var)
        logger.debug(flask_app.config)

    except Exception as e:
        print(f"{config_file_var} environment variable not defined!")
        print(e)


def construct_session_persistence_backend(flask_app):
    # A REDIS instance needs to be available. Check it
    # A local REDIS could be as simple as:
    #
    # docker run --rm -p 6379:6379 redis:alpine
    #
    d = {}
    if 'REDIS_HOST' in flask_app.config:
        r_host = flask_app.config['REDIS_HOST']
        d["SESSION_KEY_PREFIX"] = f"{app_acronym}:"
        d["SESSION_PERMANENT"] = False
        rs2 = None
        if r_host == "redis_lite":
            try:
                import redislite
                rs2 = redislite.Redis(f"tmp_{app_acronym}_backend_redislite.db")  # serverconfig={'port': '6379'}
                d["SESSION_TYPE"] = "redis"
                d["SESSION_REDIS"] = rs2
                # d["PERMANENT_SESSION_LIFETIME"] = 3600
            except ImportError as e:
                print("Package 'redislite' not found. Please, either change REDIS_HOST configuration variable to "
                      "'filesystem' or 'redis', or execute 'pip install redislite' and retry")
                sys.exit(1)
        elif r_host.startswith("filesystem:"):
            d["SESSION_TYPE"] = "filesystem"
            if flask_app.config.get("REDIS_HOST_FILESYSTEM_DIR"):
                d["SESSION_FILE_DIR"] = flask_app.config.get("REDIS_HOST_FILESYSTEM_DIR")
            d["SESSION_FILE_THRESHOLD"] = 100
            # d["SESSION_FILE_MODE"] = 666
        else:
            rs2 = redis.Redis(r_host)
            d["SESSION_TYPE"] = "redis"
            d["SESSION_REDIS"] = rs2
            # d["PERMANENT_SESSION_LIFETIME"] = 3600
        if rs2:
            try:
                print("Trying connection to REDIS '" + r_host + "'")
                rs2.ping()
                print("Connected to REDIS instance '" + r_host + "'")
            except:
                print("REDIS instance '" + r_host + "' not reachable, exiting now!")
                sys.exit(1)
        elif "SESSION_TYPE" not in d:
            print("No session persistence backend configured, exiting now!")
            sys.exit(1)
    return d


# SOCKET INITIALIZATION
def init_socket(socketio):
    SocketService(socketio)


# Initialization of tables

tm_object_type_fields = ["id", "uuid", "name"]
tm_object_types = [  # ObjectType
    (data_object_type_id["sequence"], "22bc2577-883a-4408-bba9-fcace20c0fc8", "sequence"),
    (data_object_type_id["multiple-sequence-alignment"], "e80a7d27-3ec8-4aa1-b49c-5498e0f85bee",
     "multiple-sequence-alignment"),
    (data_object_type_id["phylogenetic-tree"], "d30120f0-28df-4bca-90e4-4f0676d1c874", "phylogenetic-tree"),
    (data_object_type_id["sequence-similarity"], "7e23991b-24a0-4da1-8251-c3c3434dfb87", "sequence-similarity"),
    (data_object_type_id["dataframe"], "964f5b72-197a-4046-993b-a54e5896fd24", "dataframe"),
    (data_object_type_id["geolayer"], "b7c404e0-8b5d-4327-b830-85f8209c06e3", "geolayer"),
    (data_object_type_id["grid"], "ffb8e897-5df8-4a01-922c-7901b63ef325", "grid"),
    (data_object_type_id["geoprocess"], "ed678c7d-d4dd-4ddc-b811-a66c030ae574", "geoprocess"),  # CProcess
    (data_object_type_id["geoprocess_instance"], "f0bcbe79-89dc-42fb-b7a8-a103d361f27e", "geoprocess_instance"),
    (data_object_type_id["case_study"], "5f4c1666-e509-436b-8844-082ebe88b2b9", "case_study"),
    (data_object_type_id["view"], "4cc3b304-afb4-445b-a9da-6c5ec99aafc8", "view"),
    (data_object_type_id["dashboard"], "633a7f00-4019-4302-9067-611bea1fc934", "dashboard"),
    (data_object_type_id["sys-functions"], "ad83dcb0-e479-4a44-acf5-387b9731e8da", "sys-functions"),
    (data_object_type_id["compute-resource"], "aa97ddad-8937-4590-afc9-dc00c5601f2a", "compute-resource"),
    (data_object_type_id["algorithms"], "a4b4a7d2-732f-4db9-9a32-c57c00881eb7", "process"),  # Algorithms
    (data_object_type_id["none"], "b5371878-582a-4758-9c7c-9e536c477992", "none")  # Nulled items
]

tm_permissions_fields = ["uuid", "name", "rank"]
tm_permissions = [  # PermissionType
    ("5ac8ed2c-20d7-4905-9184-f584947719dd", "view", 1),
    ("f19ad19f-0a74-44e8-bd4e-4762404a35aa", "read", 1),
    ("feb13f20-4223-4602-a195-a3ea14615982", "export", 2),
    ("04cac7ca-a90b-4d12-a966-d8b0c77fca70", "annotate", 3),
    ("d0924822-32fa-4456-8143-0fd48da33fd7", "contribute", 3),
    ("83d837ab-01b2-4260-821b-8c4a3c52e9ab", "share", 3),
    ("91a5b4a7-3359-4eed-98df-497c42d0c3c1", "execute", 3),
    ("640de355-f58e-4446-a12a-2a99f2cfb2eb", "import", 4),  # Modify data of existing object
    ("fb1bdaec-3379-4930-b3f8-fea67d0783b7", "edit", 4),
    ("62885891-db3c-4a06-bc28-a407ffdb08a4", "permissions", 5),
    ("d3137471-84a0-4bcf-8dd8-16387ea46a30", "delete", 6)
]

tm_default_users = {  # Identities
    "0fc2b361-847c-4b2e-8fbd-533092133eef": "admin",
    "74d20b2c-5b49-462c-8d19-8a72f47b5d1b": "_anonymous",
    "91c8008e-97d5-440c-b3fc-f5a409a44768": "test_user",
    "1c8b0500-32d2-40ce-8da0-4fc772f4c3a3": "celery_user"
}

tm_default_groups = {
    "cad9f7e0-8a91-491c-9012-8900177e132b": "all-identified",  # All except anonymous (not implemented)
    "d62bd35a-b34b-43bb-a359-323d57f95f78": "cc-lab",
    "ff5d8e35-01e8-450c-b0b5-7e4487ba4811": "governance-planning",
    "1a17a8c1-2755-4bfb-b42e-c65aa53800e6": "sequence-lab",
    "3e7dab12-dd92-49cc-83eb-d4f1c0e4a62d": "endemisms",
    "86a940ec-3327-4c62-8ca3-f4fbf47a62ce": "land-planning"
}

tm_default_roles = {
    "c79b4ff7-9576-45f6-a439-551ac23c563b": "sys-admin",
    "4cc73d65-f491-4199-a52d-fc05dd45c923": "research",
    "a9f05332-ddd6-4d93-9cec-b48ef7190633": "molecular-admin",
    "ebb39daa-98d7-4cee-9596-d2ffef70f960": "geo-admin",
    "a6e1d17f-f2a3-428a-b24c-6e7b7d71fff4": "acl-admin",
    "fedc08dc-1233-4998-9d7c-2cc305fac2a1": "compute-resources-admin",
    "038f28e3-8374-4bf1-b1aa-0d63c5920294": "metadata-admin",
    "feb13f20-4223-4602-a195-a3ea14615982": "guest",
    "0156074d-acdd-43aa-8949-c7fce254ce32": "molecular-guest",
    "999d05a1-4a16-4d11-a536-b92120288a86": "geo-guest",
    "6e7ff87f-056b-46b0-b636-c7c0eb48cabe": "owner",
    "a2aef599-a34d-4290-bde5-14899b70eff1": "bioinformatic-job-executor",
    "fcc6c1e8-8c0c-46b0-9ed5-9185a427c64b": "read-molecular-api",
    "8c5a5a19-815d-46e2-8ef9-21e964fb387d": "write-molecular-api",
    "1fa175e8-28c2-4f59-a0b9-f07649ab924b": "read-geo-api",
    "ec1a58b9-5f99-454c-a3fb-85b38dd70ccc": "write-geo-api",
    "1b7d18ab-59e9-4a47-a8ef-7aab0534dfa8": "jobs-api",
    "e8244111-04dc-4697-a8f5-3468c0fbdfe5": "read-files-api",
    "9a5e8265-38c8-4607-bff2-85f5bebe6d3a": "write-files-api",
}

tm_authenticators = {  # Authenticator
    "5b7e9e40-040b-40fc-9db3-7d707fe9617f": "firebase",
    "5f32a593-306f-4b69-983c-0a5680556fae": "local",
    "15aa399f-dd58-433f-8e94-5b2222cd06c9": "local-api-key"
}

tm_job_statuses = {
    "22bc2577-883a-4408-bba9-fcace20c0fc8": "created",
    "e80a7d27-3ec8-4aa1-b49c-5498e0f85bee": "preparing_workspace",
    "763e57b7-2636-4c04-9861-d865fe0bb5ab": "exporting_to_supported_file_formats",
    "d30120f0-28df-4bca-90e4-4f0676d1c874": "transferring_data_to_resource",
    "83084df6-7ad0-45d7-b3f1-6de594c78611": "submitting",
    "a61fc587-1272-4d46-bdd0-027cde1b8a78": "waiting_for_execution",
    "600397ef-0102-486e-a6f7-d43b0f8ce4b9": "executing",
    "bfc0c9fe-631f-44d0-8e96-e22d01ffb1ed": "transferring_data_from_resource",
    "788065a7-d9f5-46fa-b8ba-8bc223d09331": "importing_into_database",
    "7e23991b-24a0-4da1-8251-c3c3434dfb87": "completed_successfully",
    "0fba3591-4ffc-4a88-977a-6e1d922f0735": "completed_error",
    "dec7e901-b3f4-4343-b3d1-4fa5fbf3804e": "cleaning_up_workspace",
    "65171b3f-e13d-4f5b-8604-2331d292baa5": "cancelling",
    "013b2f3b-4b2f-4b6c-8f5f-425132aea74b": "cancelled",
    "3eef41be-fde1-4ad4-92d0-fe795158b41d": "paused",
}

tm_job_mgmt_types = {
    "38fb34f7-a952-4036-9b0b-4d6c59e8f8d4": "galaxy",
    "0292821a-dd33-450a-bdd8-813b2b95c456": "ssh",
    "02f44e54-f139-4ea0-a1bf-fe27054c0d6c": "slurm",
    "fc1fb247-6b76-420c-9c48-f69f154cbe1d": "ebi",
    "7ece7ead-4de9-436f-a12d-86350129e444": "cipres",
}

tm_processes = {  # Preloaded processes
    #TODO "64249aae-52c6-4225-bbcf-be29286d00af": "geoprocess",
    "c8df0c20-9cd5-499b-92d4-5fb35b5a369a": "MSA ClustalW",
    "c87f58b6-cb06-4d39-a0b3-72c2705c5ae1": "PAUP Parsimony",
    "3e0240e8-b978-48a2-8fdd-9f31f4264064": "Phylogenetic Diversity Analyzer",
    "903a73a9-5a4e-4cec-b8fa-4fc9bd5ffab5": "MAFFT",
    "ea647c4e-2063-4246-bd9a-42f6a57fb9ea": "Mr Bayes",
    "985c01ca-d9d2-4df5-a8b9-8a6da251d7d4": "BLAST",
    "c55280d0-f916-4401-a1a4-bb26d8179fd7": "MSA ClustalW + PAUP Parsimony",
    "ce018826-7b20-4b70-b9b3-168c0ba46eec": "PAUP Parsimony + Phylogenetic Diversity Analyzer",
    "5b315dc5-ad12-4214-bb6a-bf013f0e4b8c": "MSA ClustalW + PAUP Parsimony + Phylogenetic Diversity Analyzer",
}

tm_system_functions = {
    "a8be2137-a6ab-4bfa-8319-6fbf0c0bb346":	"gui-blasts",
    "b7cf04d7-5860-4f81-9c8f-a7ad8212f89e":	"gui-blast-import",
    "d07ff8b7-e6e3-4a6c-8e61-8243d803e7fa":	"gui-blast-export",
    "5a2a2e8c-10ef-4fa8-816b-a837da3373cc":	"gui-blast-read",
    "af8b3d88-0527-4e37-b411-cdef86e66b81":	"gui-blast-read-content",
    "685be394-8099-48a9-bfe7-a3836fa860db":	"gui-blast-edit",
    "898dfc3f-fa3c-4364-9c01-d2b17e5adb6e":	"gui-blast-delete",
    "83937f8a-a5cc-40fb-a21c-bfdb622b3702":	"gui-blast-acl",
    "06a0606c-bc13-4cad-bebf-73cb7f409239":	"gui-collections",
    "1d0c423a-a776-4a48-9cb8-d71416e289bf":	"gui-collection-create",
    "1a610222-a907-4041-952b-30e56addc0c8":	"gui-collection-read",
    "a08fecfd-9332-426c-9348-cd69db044bb4":	"gui-collection-edit",
    "6c70fddd-8684-4da5-9d7d-ef39570e3bc2":	"gui-collection-delete",
    "6f1810a4-2985-4e37-b8cc-7635e32d9e4c":	"gui-collection-acl",
    "5953ac54-22d2-4265-b5dc-ff68d9fbaf77":	"gui-cmatrices",
    "d0034d6c-949d-4a24-b208-9f6fa1dbdd22":	"gui-cmatrix-create",
    "3440f986-46ce-4050-ab06-0471c7aaa222":	"gui-cmatrix-read",
    "77d59cf5-571f-4c90-9592-b0f56615bb15":	"gui-cmatrix-edit",
    "1deddc20-c2d0-43c2-bb1c-a5e2c4d694f3":	"gui-cmatrix-delete",
    "79df609a-f16c-4141-905e-d8d012ef89e0":	"gui-cmatrix-acl",
    "878b1caa-ef5a-45b7-9fb0-063e9900f7ac":	"gui-mas",
    "f9845e64-e5f9-4af0-9339-9e0fdc84920d":	"gui-ma-import",
    "e6986c0f-a951-44d2-b3e8-f978171682ac":	"gui-ma-export",
    "939df19e-cf0f-4c8a-b157-be3d12b161eb":	"gui-ma-read",
    "a32f91cc-6f26-4cdf-b8aa-df5ccefd3f30":	"gui-ma-read-content",
    "29958063-c8a3-4912-af38-e56148d86ee2":	"gui-ma-edit",
    "0fea25d2-9836-4e2b-a390-0d77a161bbab":	"gui-ma-acl",
    "fa1d24f9-96b6-4dc3-a97e-c1fce6b1b537":	"gui-pts",
    "e57c4588-a2d4-429f-8359-c21f3e8531e5":	"gui-pt-import",
    "d8c0481e-8d40-4dd3-92c4-075895cea3da":	"gui-pt-export",
    "4bd27b17-37f6-40a7-8290-d1eab5d60ea4":	"gui-pt-read",
    "6866a3d6-0a77-4e29-a3be-318eab5fbce2":	"gui-pt-edit",
    "3d5f4e17-ae29-47b3-a7b4-7283ab6a5114":	"gui-pt-delete",
    "305ec3bb-d940-4e25-8ef9-7854b742733b":	"gui-pt-acl",
    "06fe47dd-e52d-4e26-819f-6d8632d05609":	"gui-seqs",
    "92dd2092-77a3-4d81-8731-eebdc1e0e2fd":	"gui-seq-import",
    "884f6f07-85d5-4418-bce1-4fc8a600ff83":	"gui-seq-export",
    "cc97be39-10a4-449f-ae00-76c58aa99f6a":	"gui-seq-read",
    "243774b9-9d0f-4a01-a9e1-7359e7ccc8c2":	"gui-seq-read-content",
    "3211f86b-bb0b-4c4e-9874-611b1f9ba7ba":	"gui-seq-edit",
    "33a3903a-caf7-4167-bc05-ddab1a88e85f":	"gui-seq-delete",
    "656c2163-4a59-4eac-808f-8ffa27f075f3":	"gui-seq-acl",
    "7fc54235-1aa8-496d-939e-7d3f185f01d5":	"gui-smatrices",
    "5e558c50-f26c-44de-b626-a51d0380da5e":	"gui-smatrix-import",
    "38242707-416c-41df-9ba9-a9dba8e7fa3f":	"gui-smatrix-export",
    "f86bab48-8e9a-4ba3-a492-a0784f563e04":	"gui-smatrix-create",
    "b9bfbd98-5051-49a0-95b9-2fcf9aa71b9b":	"gui-smatrix-read",
    "5f01c6d2-23e1-4776-80cd-af58bb3c5346":	"gui-smatrix-read-content",
    "da1000e4-f535-44d3-bc45-4f4ff1f87c17":	"gui-smatrix-edit",
    "fee7ae36-b83d-4aff-98f7-296b6539302f":	"gui-smatrix-delete",
    "ec211aa4-f664-4210-924d-c17b9dc49163":	"gui-smatrix-acl",
    "2797775a-cab3-47b5-979c-2f3d9b2f2b5b":	"gui-layers",
    "442343a6-16e8-4e41-9e75-57f770945e98":	"gui-layer-import",
    "4a4f7238-0b30-45d2-bbb0-9c5f67bedda3":	"gui-layer-export",
    "d8bc90a8-a3bf-4384-8121-fbeaa2b0650f":	"gui-layer-read",
    "a203b85b-badb-4a7d-b239-49cc17b2aff2":	"gui-layer-edit",
    "eb06f829-7535-4532-a886-f5b1042ef66b":	"gui-layer-delete",
    "437de350-96d1-4a04-8c1c-7468b2f4eda4":	"gui-layer-acl",
    "fa4b4a7e-03eb-492c-bcce-cb04e6c07e14": "gui-layers-graphical",
    "dad2792c-dc80-4ba6-a6a7-9bfb7b64cd26":	"gui-jobs",
    "9982fa19-04c1-4796-8381-fd1f5fcc931f":	"gui-jobs-executing",
    "f7d5e190-1efd-4ea1-b375-66225afd3e53":	"gui-jobs-finished",
    "cd61e65d-a750-4981-a1ea-1bd994c7a77f":	"gui-job-create",
    "d49ba39d-86da-4fe1-a3fc-6b78a09f38a7":	"gui-job-read",
    "dfd8236f-63b3-44bf-b58e-e46c209c5908":	"gui-job-respawn",
    "5462a3d3-4e8b-4ca0-b0ea-a3a461d234f7":	"gui-job-cancel",
    "ffa83589-d5c5-4648-86f5-f07a5277defb":	"gui-layers-graphical",
    "682c7294-7a14-4a13-b63d-e73ba4fff0ce":	"gui-seqs-graphical",
    "d0e908f4-f0fb-4229-8f79-96fc62c24aee":	"gui-case-studies",
    "c7a38c19-a6d3-4e9d-9cd4-407063ce5e5d":	"gui-case-study-create",
    "3f2b5c80-97fd-47e8-a65e-7cd86a2e3148":	"gui-case-study-read",
    "127ca7b6-9754-4c8c-83c1-e82f30942891":	"gui-case-study-delete",
    "7ddcafe0-b735-446e-b158-5f90e3b6c546":	"gui-case-study-edit",
    "d5f6f175-f8b6-4957-983c-d2aea19478d1":	"gui-case-study-acl",
    "511402cd-07e7-44da-8508-9fc43d64e02f":	"gui-subjects",
    "02c87f97-657a-4b31-a1df-1912f1297aa1":	"gui-sources",
    "0f5d3b99-3814-4def-ad6f-1adb71e0d4de":	"gui-crs",
    "1564eebe-2101-4835-af2c-8097c49a174b":	"gui-analyses",
    "f4d03150-a416-41ac-a511-ba4af3066ee2":	"gui-analysis-create",
    "8a51c3e7-8bda-44aa-ac56-0598c5d3f43f":	"gui-analysis-read",
    "e4d6d092-3036-479d-a62b-64b253188bd5":	"gui-analysis-edit",
    "e7fc591c-f737-4fff-b31f-0c7b013be6b3":	"gui-analysis-delete",
    "6d4de88e-5081-444d-b9c0-ca8fcaee45cf":	"gui-ontologies",
    "279be182-09c4-4693-b802-24c64125a844":	"gui-ontology-import",
    "101adefe-36cb-4e49-b99f-afe687da9f3e":	"gui-ontology-read",
    "c0d2cca9-4b25-4f1b-880d-e9cfafe5771e":	"gui-ontology-edit",
    "af963780-2eff-4233-9d5a-ef40c59a2387":	"gui-ontology-delete",
    "dc01b386-29bf-48c6-b7c8-912b3ef7a4b7":	"gui-publications",
    "c81ff697-ab62-4a62-b3bb-1626dc41a0e7":	"gui-publication-create",
    "1f2cd272-5bd0-4c31-8482-2ef63814bff8":	"gui-publication-read",
    "1a3f8a6f-a025-43aa-a93e-0d8fb8caffd7":	"gui-publication-edit",
    "8008882a-db9a-4a21-93f3-8c8f6cafcc81":	"gui-publication-delete",
    "3d3d506e-07b9-4946-87be-4463add566e6":	"gui-taxonomies",
    "2a19277e-db56-47be-bfa4-8b22ea4a3fc3":	"gui-taxonomy-import",
    "747150ac-088d-49d1-b24e-4b8f2af5f6bf":	"gui-taxonomy-read",
    "4b94e284-805b-4b2d-a07b-5ca09f62d5ea":	"gui-taxonomy-edit",
    "7bf11028-068d-4d2b-8700-4435b558f2d7":	"gui-taxonomy-delete",
    "511063b9-2adb-4fa2-ab40-5dc01901ffa0":	"gui-cresources",
    "abd68dd7-6a8b-4e83-b498-5f95aad5e16d":	"gui-cresource-create",
    "021bf9eb-5737-4532-a884-b017f7c591b7":	"gui-cresource-read",
    "f26460c0-67ee-458b-b2d0-4d6f12a71ab7":	"gui-cresource-edit",
    "ae41a31e-aaf8-46fd-a205-b71ceebe8f8f":	"gui-cresource-delete",
    "b477cc87-c62a-42f7-920d-e7c6437a6d89":	"gui-cresource-acl",
    "0be2fc19-567f-4e38-9492-eaf7487bfdbd":	"gui-algorithms",
    "1ee7dfa4-294b-4c3f-9130-9ad23b46f2e9":	"gui-algorithm-create",
    "c1f4aaee-3272-4927-93d7-cc3889d2e149":	"gui-algorithm-read",
    "03bf3cb6-c607-4b0e-99a1-9f21f9e2cc86":	"gui-algorithm-edit",
    "4795a7b9-ca4a-4d57-9f98-565bd3d3ef8b":	"gui-algorithm-delete",
    "daba6f5e-ac9e-4cf2-bea1-5c21a3650946":	"gui-algorithm-acl",
    "5984dfae-f28f-4ad3-a365-a52b8e221a50":	"gui-users",
    "ee930ee9-506a-4a14-838e-5c71cd32538d":	"gui-users-authorize",
    "931a0e70-649d-40f9-97f7-482cfbc4d52b":	"gui-user-read",
    "325e40b8-09c6-4c42-9a52-fb6a63d961d9":	"gui-user-edit",
    "37905319-2abc-4ec8-aee9-f8e870953c59":	"gui-user-delete",
    "78743d6e-811d-481d-9cf1-74b9ab4c2f4f":	"gui-roles",
    "70a68ca9-01f0-40cc-a036-92496c54a4e0":	"gui-role-edit",
    "69089521-febd-466a-b6a5-1101b9711da2":	"gui-role-delete",
    "a829593d-9dc3-469d-8788-11ebb5fb811a":	"gui-organizations",
    "f5180884-87ba-4a73-aaec-92de5b29ac9a":	"gui-organization-read",
    "82667972-1cbf-4ce7-80c3-07ed23af703a":	"gui-organization-edit",
    "1557fcab-0a14-4717-95e3-b5bf191edde7":	"gui-organization-delete",
    "af324875-8d7a-4110-bf29-b4ea0459de42": "gui-groups",
    "6e228af6-2edc-408d-9c1c-546d68908274": "gui-authenticators",
    "f308f62b-d8a1-45f6-9532-12096369c8e7": "gui-permission-types",
    "880042e1-ca8a-4008-828a-e16fbb1d1a27": "gui-system-functions",
    # API
    "6ba7c06b-c164-4049-8259-f713920284a2": "get sequence",
    "7d26e0fe-501e-4568-92a9-6c250eceb705": "post sequence",
    "98da9069-5c62-44c3-8f69-985e439d106d": "put sequence",
    "a2aef599-a34d-4290-bde5-14899b70eff1": "delete sequence",
    "1a17a8c1-2755-4bfb-b42e-c65aa53800e6": "get sequences",
    "3e7dab12-dd92-49cc-83eb-d4f1c0e4a62d": "post sequences",
    "86a940ec-3327-4c62-8ca3-f4fbf47a62ce": "put sequences",
    "0b827be9-99c4-42d5-ad9a-8e7f3a4ebc39": "delete sequences",
    "dc0ae8e7-b05c-4630-95d5-fd823ca3091a": "get jobs",
    "1882a009-a285-45a4-bcd3-bb8e23bab2b6": "get job",
    "7b5e6398-3a35-400c-92cc-551687058cd0": "post job",
    "ab579882-0060-45ed-b747-5a43b39c8d25": "delete job",

}


def prepare_system_functions_rules(xlsx_file_name):
    """
    Prepare system functions rules
    """
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_file_name)
    ws = wb.get_sheet_by_name('Funciones por rol')
    system_functions_rules = {}
    cols = {3: "sys-admin",
            4: "guest",
            5: "molecular-admin",
            6: "geo-admin",
            7: "acl-admin",
            8: "compute-resources-admin",
            9: "metadata-admin",
            10: "molecular-guest",
            11: "geo-guest"
            }
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        if row[1].value is not None:
            roles = []
            for col, role in cols.items():
                if row[col].value is not None:
                    roles.append(f"'{role}'")
            if len(roles) == 0:
                roles = ["'dont-show'"]
            system_functions_rules[row[1].value] = f"role in ({','.join(roles)})"
    return "tm_system_functions_rules = {\n" + ",\n".join([f'    "{k}": "{v}"' for k, v in system_functions_rules.items()]) + "\n}"


# To update "tm_system_functions_rules" using Google Sheets:
#    https://docs.google.com/spreadsheets/d/1fXj_nF640mDUSGClA_d7AEzCCLKtmwIww9xiHWmR2qs/edit#gid=0
# ** Export to an XLSX file
# ** Uncomment the following "print" statement, and put a breakpoint in it,
# ** Then execute in Debug mode, using "main.py", "Step over" and
# ** Copy-Paste resulting string from the console, replacing the current "tm_system_functions_rules"
# print(prepare_system_functions_rules("/home/rnebot/Downloads/Lista de permisos - NEXTGENDEM.xlsx"))


tm_system_functions_rules = {
    "gui-blasts": "role in ('sys-admin','guest','molecular-admin','acl-admin')",
    "gui-blast-import": "role in ('sys-admin','molecular-admin')",
    "gui-blast-export": "role in ('sys-admin','molecular-admin')",
    "gui-blast-read": "role in ('sys-admin','guest','molecular-admin','acl-admin')",
    "gui-blast-read-content": "role in ('sys-admin','molecular-admin','molecular-guest')",
    "gui-blast-edit": "role in ('sys-admin','molecular-admin')",
    "gui-blast-delete": "role in ('sys-admin','molecular-admin')",
    "gui-blast-acl": "role in ('sys-admin','molecular-admin','acl-admin')",
    "gui-collections": "role in ('sys-admin','molecular-admin','acl-admin')",
    "gui-collection-create": "role in ('sys-admin','molecular-admin')",
    "gui-collection-read": "role in ('sys-admin','molecular-admin','acl-admin')",
    "gui-collection-edit": "role in ('sys-admin','molecular-admin')",
    "gui-collection-delete": "role in ('sys-admin','molecular-admin')",
    "gui-collection-acl": "role in ('sys-admin','molecular-admin','acl-admin')",
    "gui-cmatrices": "role in ('sys-admin','molecular-admin','acl-admin')",
    "gui-cmatrix-create": "role in ('sys-admin','molecular-admin')",
    "gui-cmatrix-read": "role in ('sys-admin','molecular-admin','acl-admin')",
    "gui-cmatrix-edit": "role in ('sys-admin','molecular-admin')",
    "gui-cmatrix-delete": "role in ('sys-admin','molecular-admin')",
    "gui-cmatrix-acl": "role in ('sys-admin','molecular-admin','acl-admin')",
    "gui-mas": "role in ('sys-admin','guest','molecular-admin','acl-admin')",
    "gui-ma-import": "role in ('sys-admin','molecular-admin')",
    "gui-ma-export": "role in ('sys-admin','molecular-admin')",
    "gui-ma-read": "role in ('sys-admin','guest','molecular-admin','acl-admin')",
    "gui-ma-read-content": "role in ('sys-admin','molecular-admin','molecular-guest')",
    "gui-ma-edit": "role in ('sys-admin','molecular-admin')",
    "gui-ma-acl": "role in ('sys-admin','molecular-admin','acl-admin')",
    "gui-pts": "role in ('sys-admin','guest','molecular-admin','acl-admin')",
    "gui-pt-import": "role in ('sys-admin','molecular-admin')",
    "gui-pt-export": "role in ('sys-admin','molecular-admin')",
    "gui-pt-read": "role in ('sys-admin','guest','molecular-admin','acl-admin')",
    "gui-pt-edit": "role in ('sys-admin','molecular-admin')",
    "gui-pt-delete": "role in ('sys-admin','molecular-admin')",
    "gui-pt-acl": "role in ('sys-admin','molecular-admin','acl-admin')",
    "gui-seqs": "role in ('sys-admin','guest','molecular-admin','acl-admin')",
    "gui-seq-import": "role in ('sys-admin','molecular-admin')",
    "gui-seq-export": "role in ('sys-admin','molecular-admin')",
    "gui-seq-read": "role in ('sys-admin','guest','molecular-admin','acl-admin')",
    "gui-seq-read-content": "role in ('sys-admin','molecular-admin','molecular-guest')",
    "gui-seq-edit": "role in ('sys-admin','molecular-admin')",
    "gui-seq-delete": "role in ('sys-admin','molecular-admin')",
    "gui-seq-acl": "role in ('sys-admin','molecular-admin','acl-admin')",
    "gui-smatrices": "role in ('sys-admin','guest','molecular-admin','acl-admin')",
    "gui-smatrix-import": "role in ('sys-admin','molecular-admin')",
    "gui-smatrix-export": "role in ('sys-admin','molecular-admin')",
    "gui-smatrix-create": "role in ('sys-admin','molecular-admin')",
    "gui-smatrix-read": "role in ('sys-admin','guest','molecular-admin','acl-admin')",
    "gui-smatrix-read-content": "role in ('sys-admin','molecular-admin','molecular-guest')",
    "gui-smatrix-edit": "role in ('sys-admin','molecular-admin')",
    "gui-smatrix-delete": "role in ('sys-admin','molecular-admin')",
    "gui-smatrix-acl": "role in ('sys-admin','molecular-admin','acl-admin')",
    "gui-layers": "role in ('sys-admin','guest','molecular-admin','geo-admin','acl-admin')",
    "gui-layer-import": "role in ('sys-admin','geo-admin')",
    "gui-layer-export": "role in ('sys-admin','geo-admin','geo-guest')",
    "gui-layer-read": "role in ('sys-admin','guest','molecular-admin','geo-admin','acl-admin')",
    "gui-layer-edit": "role in ('sys-admin','geo-admin')",
    "gui-layer-delete": "role in ('sys-admin','geo-admin')",
    "gui-layer-acl": "role in ('sys-admin','geo-admin','acl-admin')",
    "gui-layers-graphical": "role in ('sys-admin','guest','molecular-admin','geo-admin','acl-admin')",
    "gui-jobs": "role in ('sys-admin','compute-resources-admin')",
    "gui-jobs-executing": "role in ('sys-admin','compute-resources-admin')",
    "gui-jobs-finished": "role in ('sys-admin','compute-resources-admin')",
    "gui-job-create": "role in ('sys-admin','compute-resources-admin')",
    "gui-job-read": "role in ('sys-admin','compute-resources-admin')",
    "gui-job-respawn": "role in ('sys-admin','compute-resources-admin')",
    "gui-job-cancel": "role in ('sys-admin','compute-resources-admin')",
    "gui-seqs-graphical": "role in ('sys-admin','guest','molecular-admin')",
    "gui-case-studies": "role in ('sys-admin','guest','molecular-admin','acl-admin')",
    "gui-case-study-create": "role in ('sys-admin','molecular-admin','geo-admin')",
    "gui-case-study-read": "role in ('sys-admin','guest','molecular-admin','geo-admin','acl-admin')",
    "gui-case-study-delete": "role in ('sys-admin','molecular-admin','geo-admin')",
    "gui-case-study-edit": "role in ('sys-admin','molecular-admin','geo-admin')",
    "gui-case-study-acl": "role in ('sys-admin','molecular-admin','geo-admin','acl-admin')",
    "gui-subjects": "role in ('sys-admin','geo-admin','metadata-admin')",
    "gui-sources": "role in ('sys-admin','geo-admin','metadata-admin')",
    "gui-crs": "role in ('sys-admin','geo-admin','metadata-admin')",
    "gui-analyses": "role in ('sys-admin')",
    "gui-analysis-create": "role in ('sys-admin')",
    "gui-analysis-read": "role in ('sys-admin')",
    "gui-analysis-edit": "role in ('sys-admin')",
    "gui-analysis-delete": "role in ('sys-admin')",
    "gui-ontologies": "role in ('sys-admin','metadata-admin')",
    "gui-ontology-import": "role in ('sys-admin','metadata-admin')",
    "gui-ontology-read": "role in ('sys-admin','metadata-admin')",
    "gui-ontology-edit": "role in ('sys-admin','metadata-admin')",
    "gui-ontology-delete": "role in ('sys-admin','metadata-admin')",
    "gui-publications": "role in ('sys-admin')",
    "gui-publication-create": "role in ('sys-admin')",
    "gui-publication-read": "role in ('sys-admin')",
    "gui-publication-edit": "role in ('sys-admin')",
    "gui-publication-delete": "role in ('sys-admin')",
    "gui-taxonomies": "role in ('sys-admin','metadata-admin')",
    "gui-taxonomy-import": "role in ('sys-admin','metadata-admin')",
    "gui-taxonomy-read": "role in ('sys-admin','metadata-admin')",
    "gui-taxonomy-edit": "role in ('sys-admin','metadata-admin')",
    "gui-taxonomy-delete": "role in ('sys-admin','metadata-admin')",
    "gui-cresources": "role in ('sys-admin','compute-resources-admin')",
    "gui-cresource-create": "role in ('sys-admin','compute-resources-admin')",
    "gui-cresource-read": "role in ('sys-admin','compute-resources-admin')",
    "gui-cresource-edit": "role in ('sys-admin','compute-resources-admin')",
    "gui-cresource-delete": "role in ('sys-admin','compute-resources-admin')",
    "gui-cresource-acl": "role in ('sys-admin','compute-resources-admin')",
    "gui-algorithms": "role in ('sys-admin','acl-admin','compute-resources-admin')",
    "gui-algorithm-create": "role in ('sys-admin','compute-resources-admin')",
    "gui-algorithm-read": "role in ('sys-admin','acl-admin','compute-resources-admin')",
    "gui-algorithm-edit": "role in ('sys-admin','compute-resources-admin')",
    "gui-algorithm-delete": "role in ('sys-admin','compute-resources-admin')",
    "gui-algorithm-acl": "role in ('sys-admin','acl-admin','compute-resources-admin')",
    "gui-users": "role in ('sys-admin','acl-admin')",
    "gui-users-authorize": "role in ('sys-admin','acl-admin')",
    "gui-user-read": "role in ('sys-admin','acl-admin')",
    "gui-user-edit": "role in ('sys-admin','acl-admin')",
    "gui-user-delete": "role in ('sys-admin','acl-admin')",
    "gui-roles": "role in ('sys-admin','acl-admin')",
    "gui-role-edit": "role in ('sys-admin','acl-admin')",
    "gui-role-delete": "role in ('sys-admin','acl-admin')",
    "gui-organizations": "role in ('sys-admin','acl-admin')",
    "gui-organization-read": "role in ('sys-admin','acl-admin')",
    "gui-organization-edit": "role in ('sys-admin','acl-admin')",
    "gui-organization-delete": "role in ('sys-admin','acl-admin')",
    "gui-groups": "role in ('sys-admin')",
    "gui-authenticators": "role in ('sys-admin')",
    "gui-permission-types": "role in ('sys-admin')",
    "gui-system-functions": "role in ('sys-admin')"
}

tm_browser_filter_form_fields = ("id", "uuid")
tm_browser_filter_forms = [
    (data_object_type_id["sequence"], "6bb4cbbe-d6cd-4c2e-bb59-782e3c9e9f6c"),
    (data_object_type_id["multiple-sequence-alignment"], "44363784-d304-4e7e-a507-8bae48598e50"),
    (data_object_type_id["phylogenetic-tree"], "8b62f4aa-d32a-4841-89f5-9ed50da44121"),
    (data_object_type_id["geolayer"], "b4d58b86-114d-47fd-b813-ce3fcf18cbf1"),
    (data_object_type_id["grid"], "dda18cf6-0449-4824-9a6d-3b86760f27ff"),
    (data_object_type_id["dataframe"], "c645a547-49c4-49d9-a151-a81ad404c38e"),
    (100, "a2781f70-46b9-45c5-b3e7-48295d93d339"),  # Processes
    (101, "409e25a3-c2d2-468f-ac29-ab0493df2b5c"),  # Process executions
]

ht_cl_name = "Code List"
tm_hierarchy_types = {
    "45e91fea-6ca3-40c8-a798-98ca6a8e0e2a": ht_cl_name,
    "f3625ee6-99e1-4db9-82e4-adb6c1a01762": "Hierarchical Code List",
}

h_subjects_name = "Temas"
h_sources_name = "Fuentes"
h_crs_name = "CRS"
tm_hierarchies_fields = ((HierarchyType, "name", "id", "h_type_id"), "uuid", "name")
tm_hierarchies = [
    (ht_cl_name, "9975f4a0-a321-409a-bb2b-7afd05a60117", h_subjects_name),
    (ht_cl_name, "8267366d-97c7-47d8-83b5-68fe763c5e81", h_sources_name),
    (ht_cl_name, "7cdc80c7-739d-4787-8529-b52f83994930", h_crs_name)
]

tm_code_list_fields = ((Hierarchy, "name", "id", "hierarchy_id"), "uuid", "name")
tm_code_list_subjects = [
    (h_subjects_name, "2d1d71be-2e7e-4c0a-8d4c-21d2bc5bdad6", "Altimetría y modelos digitales"),
    (h_subjects_name, "eb85c308-61e3-4be6-b2c4-4d86b46a62f8", "Ámbitos de Ordenación Específicos"),
    (h_subjects_name, "68ecc77c-bc78-4352-a407-ee7bb06c04d1", "Áreas de Especial Protección"),
    (h_subjects_name, "08bd7b77-a977-4447-962b-78c0796bc06a", "Cartografía Específica"),
    (h_subjects_name, "dd1cd206-a45c-43d4-95aa-a0f1ac3c5f96", "Catastro"),
    (h_subjects_name, "1247d2cd-dcd1-4238-b44f-ea816c97b084", "Clima"),
    (h_subjects_name, "cb6b0956-4511-42d9-a75b-3b45f094a786", "Coberturas físicas y biológicas"),
    (h_subjects_name, "fdef90ae-c9c1-4415-8bf8-19b31799beb5", "Datos Estadísticos"),
    (h_subjects_name, "343e8354-925f-4c8e-aefa-7834ff2c3cc6", "Edafología"),
    (h_subjects_name, "763a0def-dc78-4561-9f8d-042d09b36ff9", "Elementos hidrográficos"),
    (h_subjects_name, "c5832d73-ba79-46e0-9448-21d0c101927f", "Entidades de población"),
    (h_subjects_name, "efc017c3-45f8-4811-9262-bfb92efc3204", "Equipamientos y servicios"),
    (h_subjects_name, "3fa8c241-05e8-40a3-8f75-e193bfe69999", "Geocodificaciones"),
    (h_subjects_name, "a26d12d3-ead3-4681-95d6-c6fb0c6cbb10", "Infraestructuras y redes de transporte"),
    (h_subjects_name, "1bfd9a77-891d-4863-aa33-09c5976ee823", "Instalaciones agrarias y acuicultura"),
    (h_subjects_name, "57acf475-21d0-4270-a947-5b20ab299539", "Instalaciones Industriales"),
    (h_subjects_name, "f5c2afb6-27ff-4118-91a5-75e6a9541763", "Malla demográfica"),
    (h_subjects_name, "42a78b57-b610-4b66-af4d-ebf090666cba", "Nomenclator de Población"),
    (h_subjects_name, "86ee31c5-bb1a-4c88-af9c-298084920c19", "Ortofotos"),
    (h_subjects_name, "d57653c7-cf75-4333-8677-c58bdf9b0a93", "Otros sin asignación"),
    (h_subjects_name, "809b6d03-db8f-4543-bc4d-93a93dca1839", "Protección Territorial"),
    (h_subjects_name, "8d746d7f-2309-4389-92d1-54ae6015a78f", "Referencias Geográficas"),
    (h_subjects_name, "0afe9d50-7ce8-41ec-9ef2-55712258c43f", "Riesgos"),
    (h_subjects_name, "1497662c-e31a-418c-ae5e-e0eb8e98888e", "Salud y seguridad humana"),
    (h_subjects_name, "ffadcde3-60bf-43e5-9aec-1b54d89370c9", "Unidades Estadísticas"),
    (h_subjects_name, "b8e58143-fe0e-440a-995b-576d2e1e9d51", "Usos del Suelo")
]

tm_code_list_crs = [
    (h_crs_name, "9c534897-2fd6-4165-b161-c0317ff59718", "EPSG:4326"),
    (h_crs_name, "9cbb027f-e532-4d81-bb66-467a06008429", "EPSG:32628"),
]

tm_code_list_sources = [
    (h_sources_name, "d77abee2-17ff-4063-8b54-a8e35647a5f0", "Grafcan"),
    (h_sources_name, "d05c68e7-5674-4b10-8cea-7b9cb8f1b798", "ISTAC"),
    (h_sources_name, "b2be8a7a-67f2-4fd1-afee-cb89164767a9", "ITC"),
    (h_sources_name, "55e9958c-2f73-47c4-afa6-37bf23c683f9", "EUROSTAT"),
]


# 071a22e8-1d90-4900-936a-0466880480d4
# faec0128-c0e3-4c2c-bbb2-bd852430eed3
# b4f90851-250e-425f-9a94-b677b593c842
# d04fc218-8dea-4938-a4d0-49e3d8174715
# 07c7512b-d7ce-4162-85c4-cda2503f290c
# 95fd8186-927c-4712-8dbc-7fe9dbef35bb
# 848b46b0-8602-42a2-a3fd-1b9be728d729
# 79668ea7-80fa-4327-a433-721a69582542
# 35313fd1-484e-48af-b6e2-e4b7464abb64
# dc50990e-f4ad-4ef3-80c8-f68fd0b1a412
# cd9e5139-8f9b-495f-955d-61577f8531a4
# 611465a5-e57e-4cf7-907f-345b5760d235
# bcaaf1ef-abbf-4b99-8638-5d2bc604998d
# d5a8c860-c6b8-49b3-9447-f75dcb650b1f
# 5b8dd6cc-493d-4994-8fe3-111c83942d0e
# 82dc3955-5125-4341-9fe1-2bc455726bee
# edfb62c9-76dc-4bc7-9282-f927a13711d6
# 94822821-e49b-4011-8188-a7dc026fd094
# 88212405-72cb-49aa-a08f-723ab115b4ec
# 6764cf5b-021a-4922-922a-04754d3c7077
# 6ca59704-dc56-4c91-951e-386802e47dca


def initialize_database_data():
    # Load base tables
    load_table(DBSession, Identity, tm_default_users)
    load_table(DBSession, Authenticator, tm_authenticators)
    load_table_extended(DBSession, ObjectType, tm_object_type_fields, tm_object_types)
    load_table(DBSession, SystemFunction, tm_system_functions)
    load_table_extended(DBSession, PermissionType, tm_permissions_fields, tm_permissions, update=True)
    load_table(DBSession, JobStatus, tm_job_statuses)
    load_table(DBSession, JobManagementType, tm_job_mgmt_types)
    load_table(DBSession, Process, tm_processes)
    load_table(DBSession, Group, tm_default_groups)
    load_table(DBSession, Role, tm_default_roles)
    load_computing_resources(DBSession)
    load_processes_in_computing_resources(DBSession)
    load_process_input_schema(DBSession)
    load_table(DBSession, HierarchyType, tm_hierarchy_types)
    load_table_extended(DBSession, Hierarchy, tm_hierarchies_fields, tm_hierarchies)
    load_table_extended(DBSession, HierarchyNode, tm_code_list_fields, tm_code_list_subjects)
    load_table_extended(DBSession, HierarchyNode, tm_code_list_fields, tm_code_list_sources)
    load_table_extended(DBSession, HierarchyNode, tm_code_list_fields, tm_code_list_crs)
    session = DBSession()
    # ACLs for system functions
    system_functions = session.query(SystemFunction).filter(SystemFunction.name.in_(tm_system_functions_rules.keys())).all()
    system_function_uuids = {f.name: f.uuid for f in system_functions}
    obj_type = session.query(ObjectType).filter(ObjectType.name == "sys-functions").first()
    for function, expression in tm_system_functions_rules.items():
        uuid = system_function_uuids[function]
        acl = session.query(ACL).filter(and_(ACL.object_type == obj_type.id, ACL.object_uuid == uuid)).first()
        if not acl:
            acl = ACL()
            acl.object_type = obj_type.id
            acl.object_uuid = uuid
            session.add(acl)
            acl_expression = ACLExpression()
            acl_expression.acl = acl
            acl_expression.expression = expression
            session.add(acl_expression)
        else:
            acl_expression = session.query(ACLExpression).filter(ACLExpression.acl == acl).first()
            if not acl_expression:
                acl_expression = ACLExpression()
                acl_expression.acl = acl
                acl_expression.expression = expression
                session.add(acl_expression)
            else:
                acl_expression.expression = expression
    # load_table_extended(DBSession, BrowserFilterForm, tm_browser_filter_form_fields, tm_browser_filter_forms)

    # Load default authentication for "test_user"
    test_user_id = "test_user"
    authenticator_id = "5f32a593-306f-4b69-983c-0a5680556fae"  # "local". Just the user name is good to be authorized
    iden = session.query(Identity).filter(Identity.name == test_user_id).first()
    authentication = session.query(Authenticator).filter(Authenticator.uuid == authenticator_id).first()
    iden_authentication = session.query(IdentityAuthenticator). \
        filter(
        and_(IdentityAuthenticator.identity == iden, IdentityAuthenticator.authenticator == authentication)).first()
    if not iden_authentication:
        iden_authentication = IdentityAuthenticator()
        iden_authentication.identity = iden
        iden_authentication.authenticator = authentication
        iden_authentication.name = test_user_id
        iden_authentication.email = "test@test.org"
        session.add(iden_authentication)

    # Celery user, also with local authentication
    celery_user_id = "celery_user"
    iden = session.query(Identity).filter(Identity.name == celery_user_id).first()
    iden_authentication = session.query(IdentityAuthenticator). \
        filter(
        and_(IdentityAuthenticator.identity == iden, IdentityAuthenticator.authenticator == authentication)).first()
    if not iden_authentication:
        iden_authentication = IdentityAuthenticator()
        iden_authentication.identity = iden
        iden_authentication.authenticator = authentication
        iden_authentication.name = celery_user_id
        iden_authentication.email = "celery@celery.org"
        session.add(iden_authentication)

    # _anonymous user, also with local authentication
    anonymous_user_id = "_anonymous"
    iden = session.query(Identity).filter(Identity.name == anonymous_user_id).first()
    iden_authentication = session.query(IdentityAuthenticator). \
        filter(
        and_(IdentityAuthenticator.identity == iden, IdentityAuthenticator.authenticator == authentication)).first()
    if not iden_authentication:
        iden_authentication = IdentityAuthenticator()
        iden_authentication.identity = iden
        iden_authentication.authenticator = authentication
        iden_authentication.name = anonymous_user_id
        iden_authentication.email = "_@anonymous.org"
        session.add(iden_authentication)

    session.commit()

    # Set test_user roles and groups
    load_many_to_many_table(DBSession, RoleIdentity, Role, Identity, ["role_id", "identity_id"],
                            [("sys-admin", test_user_id),
                             ("sys-admin", celery_user_id),
                             ("guest", anonymous_user_id)])
    load_many_to_many_table(DBSession, GroupIdentity, Group, Identity, ["group_id", "identity_id"],
                            [("all-identified", test_user_id)])

    # Associate permission types to object types
    load_many_to_many_table(DBSession, ObjectTypePermissionType, ObjectType, PermissionType,
                            ["object_type_id", "permission_type_id"],
                            [["sequence", "read"],
                             ["sequence", "annotate"],
                             ["sequence", "delete"],
                             ["multiple-sequence-alignment", "read"],
                             ["multiple-sequence-alignment", "annotate"],
                             ["multiple-sequence-alignment", "delete"],
                             ["phylogenetic-tree", "read"],
                             ["phylogenetic-tree", "annotate"],
                             ["phylogenetic-tree", "delete"],
                             ["process", "read"],
                             ["process", "execute"],
                             ["geolayer", "view"],
                             ["geolayer", "read"],
                             ["geolayer", "export"],
                             ["geolayer", "edit"],
                             ["geolayer", "delete"],
                             ["geolayer", "permissions"],
                             ])


def initialize_database(flask_app):
    recreate_db = False
    if 'DB_CONNECTION_STRING' in flask_app.config:
        db_connection_string = flask_app.config['DB_CONNECTION_STRING']
        print(f"Connecting to {app_acronym.upper()} database server")
        print(db_connection_string)
        print("-----------------------------")
        if db_connection_string.startswith("sqlite://"):
            base_app_pkg.engine = sqlalchemy.create_engine(db_connection_string,
                                                           echo=True,
                                                           connect_args={'check_same_thread': False},
                                                           poolclass=StaticPool)
        else:
            base_app_pkg.engine = create_pg_database_engine(db_connection_string, app_acronym, recreate_db=recreate_db)

        # global DBSession # global DBSession registry to get the scoped_session
        DBSession.configure(bind=base_app_pkg.engine)  # reconfigure the sessionmaker used by this scoped_session
        orm.configure_mappers()  # Important for SQLAlchemy-Continuum
        tables = ORMBase.metadata.tables
        connection = base_app_pkg.engine.connect()
        table_existence = [base_app_pkg.engine.dialect.has_table(connection, tables[t].name) for t in tables]
        connection.close()
        if False in table_existence:
            ORMBase.metadata.bind = base_app_pkg.engine
            ORMBase.metadata.create_all()
        # connection = biobarcoding.engine.connect()
        # table_existence = [biobarcoding.engine.dialect.has_table(connection, tables[t].name) for t in tables]
        # connection.close()

        # Load base tables
        initialize_database_data()
        update_functional_object_tsvector(DBSession())
    else:
        print("No database connection defined (DB_CONNECTION_STRING), exiting now!")
        sys.exit(1)


def initialize_postgis(flask_app):
    recreate_db = False
    if 'POSTGIS_CONNECTION_STRING' in flask_app.config:
        db_connection_string = flask_app.config['POSTGIS_CONNECTION_STRING']
        print(f"Connecting to {app_acronym}_geoserver database server")
        print(db_connection_string)
        print("-----------------------------")

        base_app_pkg.postgis_engine = create_pg_database_engine(db_connection_string, f"{app_acronym}_geoserver",
                                                                recreate_db=recreate_db)
        # global DBSession # global DBSession registry to get the scoped_session
        DBSessionGeo.configure(
            bind=base_app_pkg.postgis_engine)  # reconfigure the sessionmaker used by this scoped_session
        orm.configure_mappers()  # Important for SQLAlchemy-Continuum
        connection = base_app_pkg.postgis_engine.connect()
        try:
            connection.execute("CREATE EXTENSION postgis")
        except:
            pass
        connection.execute("commit")
        connection.close()
        tables = ORMBaseGeo.metadata.tables
        connection = base_app_pkg.postgis_engine.connect()
        table_existence = [base_app_pkg.postgis_engine.dialect.has_table(connection, tables[t].name) for t in tables]
        connection.close()
        if False in table_existence:
            ORMBaseGeo.metadata.bind = base_app_pkg.postgis_engine
            ORMBaseGeo.metadata.create_all()
        # Load base tables
        # initialize_gnd_geoserver_data() # not implemented
    else:
        print("No database connection defined (POSTGIS_CONNECTION_STRING), exiting now!")
        sys.exit(1)


def register_api(bp: Blueprint, view, endpoint: str, url: str, pk='id', pk_type='int'):
    view_func = view.as_view(endpoint)
    bp.add_url_rule(url, defaults={pk: None}, view_func=view_func, methods=['GET'])
    bp.add_url_rule(url, view_func=view_func, methods=['POST'])
    bp.add_url_rule(f'{url}<{pk_type}:{pk}>', view_func=view_func, methods=['GET', 'PUT', 'DELETE'])
    return view_func


def make_simple_rest_crud(entity, entity_name: str, execution_rules: Dict[str, str] = {},
                          aux_filter=None, default_filter=None, alt_getters: Dict[str, Callable] = None):
    """
    Create a CRUD RESTful endpoint

    :param entity: the entity class to manage
    :param entity_name: the name of the entity, in plural
    :param execution_rules: a dictionary of execution rules (according to the decorator "n_session") by CRUD method
    :param aux_filter: a function to augment the filtering capabilities (applies to GET without "id_")
    :param default_filter: A string with a default "filter" condition (applies to GET without "id_")
    :param alt_getters: a dictionary of alternative internal getters
    :return: the blueprint (to register it) and the class (if needed)
    """

    class CrudAPI(MethodView):

        @n_session(read_only=True, authr=execution_rules.get("r"))
        def get(self, _id=None):  # List or Read
            db = g.n_session.db_session
            r = ResponseObject()

            if alt_getters and "get" in alt_getters:
                entities = alt_getters["get"](db, _id)
                if isinstance(entities, list):
                    r.content = entities
                    r.count = len(entities) if entities else 0
                elif isinstance(entities, entity):
                    r.content = entities
                    r.count = 1
            # The getter can return None to indicate to proceed with the default behavior, below
            # Of course, if no alt_getter was specified, also proceed with the default behavior
            if r.content is None:
                if _id is None:
                    # List of all
                    kwargs = parse_request_params()
                    from ..services import get_query
                    query, count = get_query(db, entity, aux_filter=aux_filter, default_filter=default_filter, **kwargs)
                    # TODO Detail of fields
                    r.count = count
                    r.content = query.all()
                else:
                    # Detail
                    # TODO Detail of fields
                    r.content = db.query(entity).filter(entity.id == int(_id)).first()
                    r.count = 1

            return r.get_response()

        @n_session(authr=execution_rules.get("c"))
        def post(self):  # Create
            db = g.n_session.db_session
            r = ResponseObject()
            t = request.json
            s = entity.Schema().load(t, instance=entity(), partial=True)
            db.add(s)
            db.flush()
            r.content = s
            return r.get_response()

        @n_session(authr=execution_rules.get("u"))
        def put(self, _id):  # Update (total or partial)
            db = g.n_session.db_session
            r = ResponseObject()
            t = request.json
            if alt_getters and "put" in alt_getters:
                s = alt_getters["put"](db, _id)
            else:
                s = db.query(entity).filter(entity.id == int(_id)).first()
            s = entity.Schema().load(t, instance=s, partial=True)
            db.add(s)
            db.flush()
            r.content = s
            return r.get_response()

        @n_session(authr=execution_rules.get("d"))
        def delete(self, _id):  # Delete
            db = g.n_session.db_session
            r = ResponseObject()
            # TODO Multiple delete
            # if _id is None:
            #     # List of all
            #     query = db.query(entity)
            #     # Filter, Order, Pagination
            #     kwargs = check_request_params()
            #     if 'filter' in kwargs:
            #         query = query.filter(filter_parse(entity, kwargs.get('filter'))).delete()
            # else:
            if alt_getters and "delete" in alt_getters:
                s = alt_getters["delete"](db, _id)
            else:
                s = db.query(entity).filter(entity.id == int(_id)).first()
            db.delete(s)
            return r.get_response()

    # If the following line is uncommented, it complains on "overwriting an existing endpoint function".
    # This function is public, because it is just the schema, so, no problem.
    # @n_session()
    def get_entities_json_schema():
        r = ResponseObject()
        factory = SchemaFactory(StructuralWalker)
        r.content = factory(entity)

        return r.get_response()

    @n_session()
    def get_entity_json_schema(_id):
        db = g.n_session.db_session
        r = ResponseObject()
        factory = SchemaFactory(StructuralWalker)
        ent = db.query(entity).filter(entity.id == _id).first()
        r.content = dict(schema=factory(entity), data=ent)
        return r.get_response()

    # make_simple_rest_crud(entity, entity_name: str, execution_rules: Dict[str, str] = {})
    bp_entity = Blueprint(f'bp_{entity_name}', __name__)
    register_api(bp_entity, CrudAPI, entity_name, f"{app_api_base}/{entity_name}/", "_id", "string")
    # Two equivalent endpoints to obtain the JSON schema of an entity !!
    # (useful to automatically generate a Formly form)
    bp_entity.add_url_rule(f"{app_api_base}/{entity_name}.schema.json", view_func=get_entities_json_schema,
                           methods=['GET'])
    bp_entity.add_url_rule(f"{app_api_base}/{entity_name}/<int:_id>.schema.json", view_func=get_entity_json_schema,
                           methods=['GET'])

    return bp_entity, CrudAPI


# GENERIC REST FUNCTIONS

def chew_data(param):
    try:
        param = param.decode()
    except:
        pass
    try:
        import ast
        param = ast.literal_eval(param)
    except:
        pass
    try:
        param = unquote(param)
    except:
        pass
    try:
        param = json.loads(param)
    except:
        pass
    return param


def decode_request_params(data):
    res = {}
    params = chew_data(data) or []
    if isinstance(params, (list, tuple)):
        for obj in params:
            for key in obj:
                obj[key] = chew_data(obj[key])
        return params
    else:
        for key in params:
            res[key] = chew_data(params[key])
    return res


def parse_request_params(data=None, default_kwargs=None):
    kwargs = {'filter': [], 'order': [], 'pagination': {}, 'values': {}, 'searchValue': ''}
    if not data and request:
        if request.json:
            kwargs.update(parse_request_params(request.json))
        if request.data:
            kwargs.update(parse_request_params(request.data))
        if request.values:
            kwargs.update(parse_request_params(request.values))
    elif data:
        print(f'RAW_DATA: {data}')
        input = decode_request_params(data)
        for key in ('filter', 'order', 'pagination', 'values', 'searchValue'):
            try:
                i = input.pop(key)
            except:
                continue
            kwargs[key] = i if i else kwargs[key]
        if input:
            if type(kwargs.get('values')) != type(input):
                kwargs['values'] = input
            elif isinstance(input, dict):
                kwargs['values'].update(input)
            elif isinstance(input, (list,tuple)):
                kwargs['values'] += input
    if isinstance(default_kwargs, dict):
        for k, v in default_kwargs.items():
            if k in kwargs:
                v = kwargs[k]
                if v is None or (isinstance(v, list) and len(v) == 0) or (isinstance(v, str) and v == ''):
                    kwargs[k] = default_kwargs[k]

    print(f'CLEAN_DATA: {kwargs}')
    return kwargs


def auth_filter(orm,
                permission_types_ids: Union[int, List[int]],
                object_types_ids,
                identity_id: Optional[int] = None,
                object_uuids: Optional[List[str]] = None,
                time=None,
                permission_flag=False,
                authorizable_flag=False,
                reference_entity: Union[int, str] = -1) -> Query:
    """
    !!!! ACL filter !!!!

    @param orm: Class of SQLAlchemy ORM to check. "FunctionalObject" for any class
    @param permission_types_ids: List of permission type ids (or just one) to pass the filter against
    @param object_types_ids: List of object type ids (similar to "orm")
    @param identity_id: Identity id of who is being authorized. If None, it is the user logged in the current session
    @param object_uuids: List of specific object uuids to reduce the search for authorizations
    @param time: Time, to check validity of ACL rules
    @param permission_flag: If True, return the permission type enabling access
    @param authorizable_flag: If True, return the authorizable (identity, group, organization, role) enabling access
    @param reference_entity: Reference entity id (or uuid) to check against. If -1, find default reference entity

    CollectionDetail (cd) <> Collection (c) > ACL <> ACLDetail (ad)

        SELECT
          CASE cd.object_uuid WHEN NULL THEN acl.object_uuid ELSE cd.object_uuid as oid,
          ad.permission_id as pid,
          ad.authr_id as aid  # To explain why it was authorized
        FROM CollectionDetail cd JOIN Collection c ON cd.collection_id=c.id RIGHT JOIN ACL ON c.uuid=acl.object_uuid JOIN ACLDetail ad ON acl.id=ad.acl_id
        WHERE ad.authorizable IN (<identities>)  # Authorizable
              AND object_type IN (<collection object type>[, <target object type>])  # Object types
              AND date time BETWEEN ad.validity_start AND ad.validity_end
              [AND permission_types IN (...)]
              [AND object_uuids IN (...)]

        * Añadir lista object_ids objeto sirve para ver permisos de esos objetos concretos
        * Añadir lista id tipos permisos sirve para ver si esos tipos están

    @return: <orm_clause_filter> || object_uuids[, permissions][, authorizables]
    """

    def related_authr_ids(identity_id: int):
        # Get the organizations, groups, and roles associated to an identity
        ids = DBSession.query(Authorizable.id) \
            .join(OrganizationIdentity, OrganizationIdentity.organization_id == Authorizable.id, isouter=True) \
            .join(GroupIdentity, GroupIdentity.group_id == Authorizable.id, isouter=True) \
            .join(RoleIdentity, RoleIdentity.role_id == Authorizable.id, isouter=True) \
            .filter(or_(Authorizable.id == identity_id,
                        OrganizationIdentity.identity_id == identity_id,
                        GroupIdentity.identity_id == identity_id,
                        RoleIdentity.identity_id == identity_id))
        return [i for i, in ids]

    def related_perm_ids(permission_id: int):
        # Get the mayor permissions that also allow permission_id
        ids = DBSession.query(PermissionType.id) \
            .filter(or_(PermissionType.id == permission_id,
                        PermissionType.rank >=
                        DBSession.query(PermissionType.rank).filter(PermissionType.id == permission_id)))
        return [i for i, in ids]

    def default_reference_entity(clazz, identity_id: int):
        ent = DBSession.query(clazz).filter(clazz.authr_reference == True).one_or_none()
        if ent:
            return ent.uuid
        else:
            return None

    # --------------------------- auth_filter -------------------------------------------------------------------------
    from flask import current_app
    if current_app.config["ACL_ENABLED"] != 'True':  # Disable ACL control
        return True

    identity_id = g.n_session.identity.id if identity_id is None else identity_id
    sys_admin_id = DBSession.query(Role.id).filter(Role.name == 'sys-admin').first()[0]
    if isinstance(identity_id, int):
        authorizables = related_authr_ids(identity_id)
        is_sys_admin = sys_admin_id in authorizables
    else:
        is_sys_admin = sys_admin_id in identity_id

    if is_sys_admin:  # If user has "sys-admin" role, access to everything
        return True

    # If reference entity enables access -> return True, if not, continue evaluation
    if reference_entity is not None:  # Already checked
        if reference_entity == -1:
            # Find UUID of a reference object
            reference_entity = default_reference_entity(orm, identity_id)
        if reference_entity:
            _ = auth_filter(orm,
                            permission_types_ids,
                            object_types_ids,
                            identity_id,
                            [reference_entity],
                            time,
                            False,
                            False,
                            None)
            if str(reference_entity) in set([str(i.uuid) for i in DBSession.query(orm.uuid).filter(_).all()]):
                return True

    # ACL Detail entry in valid date range
    from datetime import datetime
    time = time if time else datetime.now()
    filter_clause = [
        or_(time >= ACLDetail.validity_start, ACLDetail.validity_start == None),
        or_(time <= ACLDetail.validity_end, ACLDetail.validity_end == None)
    ]

    # ACL applies to object types (if specified)
    filter_clause.append(ACL.object_type.in_(object_types_ids)) if object_types_ids else None
    # ACL applies to specific objects (if specified)
    filter_clause.append(ACL.object_uuid.in_(object_uuids)) if object_uuids else None
    try:
        # By identity or authorizables (role, group, organization) associated with the identity
        if isinstance(identity_id, int):
            filter_clause.append(ACLDetail.authorizable_id.in_(authorizables))
        elif isinstance(identity_id, (tuple, list, set)):
            filter_clause.append(ACLDetail.authorizable_id.in_(identity_id))
        # By permission or superior permissions
        if isinstance(permission_types_ids, int):
            filter_clause.append(ACLDetail.permission_id.in_(related_perm_ids(permission_types_ids)))
        elif isinstance(permission_types_ids, (tuple, list, set)):
            filter_clause.append(ACLDetail.permission_id.in_(permission_types_ids))

        # Object in authorized collection
        collected = DBSession.query(CollectionDetail.object_uuid) \
            .join(Collection) \
            .join(ACL, Collection.uuid == ACL.object_uuid) \
            .join(ACLDetail) \
            .filter(*filter_clause)

        # Direct object
        uncollected = DBSession.query(ACL.object_uuid).join(ACLDetail).filter(*filter_clause)

        final_query = collected.union(uncollected)
        if permission_flag or authorizable_flag:
            entities = []
            entities += [ACLDetail.permission] if permission_flag else []
            entities += [ACLDetail.authorizable] if authorizable_flag else []
            q = final_query.with_entities(entities)
        else:
            q = orm.uuid.in_(final_query)

        # OWNER has unlimited permissions (no need to check ACL)
        if isinstance(identity_id, int):
            q = or_(q, orm.owner_id == identity_id)
        return q
    except Exception as e:
        print(e)
        return None


def filter_parse(orm, filter, aux_filter=None, session=None):
    """
    @param orm: <ORMSqlAlchemy>
    @param filter:
        [{"field1": "<condition>", "field2": "<condition>"}, {"field1": "<condition"}]
        <condition>: {"op": "<operador", "left": "<valor>", "right": <valor>, "unary": <valor>}
    @param aux_filter: <callable function(filter)>
    @return: <orm_clause_filter>
    """

    def get_condition(orm, field, condition):
        obj = getattr(orm, field)

        if not isinstance(condition, dict):
            if isinstance(condition, (list, tuple)):
                return obj.in_(condition)
            return obj == condition

        op = condition["op"]
        value, left, right = condition.get("unary"), condition.get("left"), condition.get("right")
        if op == "out":
            return obj.notin_(value)
        if op == "in":
            return obj.in_(value)
        if op == "eq":
            return obj == value
        if op == "between":
            return obj.between_(left, right)
        if op == "le":
            return obj <= value
        if op == "ge":
            return obj >= value
        if op == "contains":
            return value.in_(obj)

        return True

    try:
        if not isinstance(filter, (list, tuple)):
            filter = [filter]
        or_clause = []
        for clause in filter:
            and_clause = []
            for field, condition in clause.items():
                if hasattr(orm, field):
                    and_clause.append(get_condition(orm, field, condition))
                else:
                    print(f'Warning: Unknown column "{field}" for the given tables.')
            if aux_filter:
                try:
                    lst = aux_filter(clause, session)
                except:
                    lst = aux_filter(clause)
                and_clause += lst

            or_clause.append(and_(*and_clause))
        return or_(*or_clause)
    except Exception as e:
        traceback.print_exc()
        return None


def order_parse(orm, sort, aux_order=None):
    """
    @param orm: <ORMSqlAlchemy>
    @param sort: [{"order": "<asc|desc>", "field": "<name>"}]
    @param aux_order: <callable function(order)>
    @return: <orm_clause_order>
    """

    def get_condition(orm, clause):
        obj = getattr(orm, clause["field"])
        if clause["order"].startswith("asc"):
            return obj.asc()
        if clause["order"].startswith("desc"):
            return obj.desc()
        return True

    try:
        if not isinstance(sort, (list, tuple)):
            order = [sort]
        else:
            order = sort
        ord_clause = []
        for clause in order:
            if hasattr(orm, clause.get("field", "")):
                ord_clause.append(get_condition(orm, clause))
            else:
                print(f'Unknown column "{clause.get("field", "")}" for the given tables.')
                if aux_order:
                    ord_clause += aux_order(clause)
        return ord_clause
    except Exception as e:
        print(e)
        return None


def get_content(session, clazz, issues, id_=None, aux_filter=None, query=None):
    content = None
    count = 0
    try:
        if id_ is None:
            # List of all
            kwargs = parse_request_params(default_kwargs=dict(order=[dict(field="name", order="ascend")]))
            from ..services import get_query
            query2, count = get_query(session, clazz, query, aux_filter=aux_filter, **kwargs)
            content = query2.all()
        else:
            # Detail
            if not query:
                query = session.query(clazz)
            content = query.filter(clazz.id == id_).first()
    except Exception as e:
        traceback.print_exc(e)
        _, status = issues.append(Issue(IType.ERROR,
                                        f'READ "{clazz.__name__}" data: The "{clazz.__name__}"'
                                        f' data could not be read.')), 500
    if not content:
        _, status = issues.append(Issue(IType.INFO, f'no data available')), 200
    else:
        _, status = issues.append(
            Issue(IType.INFO, f'READ "{clazz.__name__}": The "{clazz.__name__}"'
                              f' data were successfully read')), 200
    return issues, content, count, status


def initialize_ssh(flask_app):
    """
    ssh_resources = conn.execute("SELECT * FROM jobs_compute_resources join jobs_job_mgmt_types as jt" +
                                 " on jm_type_id=jt.id where jt.name='ssh'").fetchall()
    :param flask_app:
    :return:
    """
    session = DBSession()
    ssh_resources = session.query(ComputeResource, JobManagementType).filter(
        ComputeResource.jm_type_id == JobManagementType.id, JobManagementType.name == 'ssh').all()
    for ssh_res in ssh_resources:
        compute_resource = ssh_res.ComputeResource
        ssh_credentials = compute_resource.jm_credentials
        ssh_location = compute_resource.jm_location
        cmd = ["ssh", "-o", "StrictHostKeyChecking=no",
               "-o", f"UserKnownHostsFile={ssh_credentials['known_hosts_filepath']}",
               f"{ssh_credentials['username']}@{ssh_location['host']}", "'echo'"]
        subprocess.run(cmd)


# GALAXY INIZIALIZATION
def install_galaxy_tools(gi, tools):
    '''
    tool_shed_url, name, owner,
    changeset_revision,
    install_tool_dependencies = False,
    install_repository_dependencies = False,
    install_resolver_dependencies = False,
    tool_panel_section_id = None,
    new_tool_panel_section_label = Non
    '''
    if isinstance(tools, list):
        for tool in tools:
            tool_shed_url = 'https://' + tool['tool_shed']
            gi.toolshed.install_repository_revision(tool_shed_url=tool_shed_url,
                                                    name=tool['name'],
                                                    owner=tool['owner'],
                                                    changeset_revision=tool['changeset_revision'],
                                                    install_tool_dependencies=True,
                                                    install_repository_dependencies=True,
                                                    install_resolver_dependencies=True,
                                                    new_tool_panel_section_label='New'
                                                    )


def check_galaxy_tools(wf1_dic, wf2_dic):
    steps1 = wf1_dic['steps']
    steps2 = wf2_dic['steps']
    tool_list = list()
    for step, content in steps1.items():
        if 'errors' in content:
            if content['errors'] == "Tool is not installed":
                # TODO depende de la versión de galaxy esto lleva un punto al final o no xq lo que hay que buscar
                #  otra cosa
                tool_list.append(steps2[step]['tool_shed_repository'])
    if len(tool_list) == 0:
        return 'all tools are installed'
    else:
        return tool_list


def initialize_galaxy(flask_app):
    session = DBSession()
    galaxy_resources = session.query(ComputeResource, JobManagementType).filter(
        ComputeResource.jm_type_id == JobManagementType.id, JobManagementType.name == 'galaxy').all()
    for galaxy_res in galaxy_resources:
        compute_resource = galaxy_res.ComputeResource
        url = compute_resource.jm_location['url']
        api_key = compute_resource.jm_credentials['api_key']

        # Install basic workflow if it is not installed
        gi = galaxy.GalaxyInstance(url=url, key=api_key)
        path = ROOT + '/biobarcoding/workflows/'
        for workflow in os.listdir(path):
            workflow_path = path + workflow
            with open(workflow_path, 'r') as f:
                wf_dict_in = json.load(f)
            name = wf_dict_in['name']
            wf = gi.workflows.get_workflows(name=name)
            w_id = wf[0].get('id') if wf else None
            if w_id:
                gi.workflows.delete_workflow(w_id)
            wf = gi.workflows.import_workflow_from_local_path(workflow_path)
            wf_dict_out = gi.workflows.export_workflow_dict(wf['id'])
            list_of_tools = check_galaxy_tools(wf_dict_out, wf_dict_in)
            install_galaxy_tools(gi, list_of_tools)
        # conversion of workflows galaxy into workflows formly
        # TODO convert_workflows_to_formly()


def initialize_database_chado(flask_app):
    cfg = flask_app.config
    if 'CHADO_CONNECTION_STRING' in flask_app.config:
        db_connection_string = cfg["CHADO_CONNECTION_STRING"]
    elif 'CHADO_DATABASE' in flask_app.config:
        db_connection_string = f'postgresql://{cfg["CHADO_USER"]}:{cfg["CHADO_PASSWORD"]}@{cfg["CHADO_HOST"]}:{cfg["CHADO_PORT"]}/{cfg["CHADO_DATABASE"]}'
        print("Connecting to Chado database server")
        print(db_connection_string)
        print("-----------------------------")
        base_app_pkg.chado_engine = sqlalchemy.create_engine(db_connection_string, echo=False)
        # global DBSessionChado # global DBSessionChado registry to get the scoped_session
        DBSessionChado.configure(
            bind=base_app_pkg.chado_engine)  # reconfigure the sessionmaker used by this scoped_session
        orm.configure_mappers()  # Important for SQLAlchemy-Continuum
        ORMBaseChado.metadata.bind = base_app_pkg.chado_engine
        ORMBaseChado.metadata.reflect()
    else:
        print("No CHADO connection defined (CHADO_CONNECTION_STRING), exiting now!")
        print(flask_app.config)
        sys.exit(1)


def initialize_chado_edam(flask_app):
    from sqlalchemy import text, Table, Index, MetaData
    cfg = flask_app.config
    if 'CHADO_DATABASE' in flask_app.config:
        chado_xml_folder = os.path.join(os.sep, "app", "docker_init", "chado_xml")
        db_relationship_comment = """Specifies relationships between databases.  This is 
        particularly useful for ontologies that use multiple prefix IDs for it\'s vocabularies. For example,
        the EDAM ontology uses the prefixes \"data\", \"format\", \"operation\" and others. Each of these would
        have a record in the db table.  An \"EDAM\" record could be added for the entire ontology to the
        db table and the previous records could be linked as \"part_of\" EDAM.  As another example
        databases housing cross-references may have sub databases such as NCBI (e.g. Taxonomy, SRA, etc).
        This table can use a \'part_of\' record to link all of them to NCBI."""

        db_connection_string = f'postgresql://{cfg["CHADO_USER"]}:{cfg["CHADO_PASSWORD"]}@{cfg["CHADO_HOST"]}:{cfg["CHADO_PORT"]}/{cfg["CHADO_DATABASE"]}'
        print("Connecting to Chado database server to insert EDAM")
        print(db_connection_string)
        print("-----------------------------")
        base_app_pkg.chado_engine = sqlalchemy.create_engine(db_connection_string, echo=True)

        with base_app_pkg.chado_engine.connect() as conn:
            relationship_id = conn.execute(
                text("select * from INFORMATION_SCHEMA.TABLES where TABLE_NAME = 'db_relationship'")).fetchone()
            if relationship_id is None:
                try:
                    meta = MetaData()
                    meta.reflect(base_app_pkg.chado_engine, only=["db",
                                                                  "cvterm"])  # this get information about the db and cvterm tables in the existing database
                    # CREATE db_relationship table to relate EDAM submodules with EDAM itself
                    db_relationship = Table('db_relationship', meta,
                                            Column('db_relationship_id', BigInteger, primary_key=True),
                                            Column('type_id',
                                                   BigInteger,
                                                   ForeignKey('cvterm.cvterm_id', ondelete="CASCADE",
                                                              initially="DEFERRED"),
                                                   nullable=False),
                                            Column('subject_id',
                                                   BigInteger,
                                                   ForeignKey('db.db_id', ondelete="CASCADE", initially="DEFERRED"),
                                                   nullable=False),
                                            Column('object_id',
                                                   BigInteger,
                                                   ForeignKey('db.db_id', ondelete="CASCADE", initially="DEFERRED"),
                                                   nullable=False),
                                            UniqueConstraint('type_id', 'subject_id', 'object_id',
                                                             name='db_relationship_c1'),
                                            Index('db_relationship_idx1', 'type_id'),
                                            Index('db_relationship_idx2', 'subject_id'),
                                            Index('db_relationship_idx3', 'object_id'),
                                            comment=db_relationship_comment
                                            )

                    db_relationship.create(bind=conn)

                except Exception as e:
                    print("Something went wrong when inserting EDAM:")
                    print(e)
            else:
                print("The preprocess to insert EDAM was already executed")

            try:
                edam_id = conn.execute(text("select db_id from db where name=\'EDAM\'")).fetchone()
                if edam_id is None:
                    # INSERT THE CHADO XML (EDAM) IN THE DATABASE
                    command = ["stag-storenode.pl", "-d",
                               f'dbi:Pg:dbname={cfg["CHADO_DATABASE"]};host={cfg["CHADO_HOST"]};port={cfg["CHADO_PORT"]}',
                               "--user", cfg["CHADO_USER"], "--password", cfg["CHADO_PASSWORD"],
                               os.path.join(chado_xml_folder, "EDAM.chado")]

                    subprocess.run(command)
                    # We commit 2 times because we need a commit on the last block to execute this one

                    # FILL THE db_relationship TABLE
                    edam_id = conn.execute(text("select db_id from db where name=\'EDAM\'")).fetchone()[0]
                    db_ids = conn.execute(text("select db_id from db where db_id > " + str(edam_id))).fetchall()
                    type_id = conn.execute(text(
                        "select cvterm_id from cvterm join cv on cvterm.cv_id = cv.cv_id where cvterm.name = \'part_of\' "
                        "and cv.name = \'relationship\'")).fetchone()[0]

                    for db_id in db_ids:
                        conn.execute(text("INSERT INTO db_relationship (type_id,subject_id,object_id)" +
                                          "VALUES(" + str(type_id) + "," + str(db_id[0]) + "," + str(
                            edam_id) + ")"))
                else:
                    print("EDAM was already inserted")

            except Exception as e:
                print("""WARNING! Something went wrong when inserting EDAM. 
                        You will need to delete the database if you do not want to insert EDAM:""")
                print(e)

    else:
        print("No CHADO connection defined (CHADO_CONNECTION_STRING), exiting now!")
        print(flask_app.config)
        sys.exit(1)
